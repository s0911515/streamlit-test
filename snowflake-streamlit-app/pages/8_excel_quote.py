import streamlit as st
import io
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

st.set_page_config(page_title="Excel帳票出力", page_icon="📗", layout="wide")

st.title("📗 Excel帳票出力（見積書）")
st.write(
    "入力した明細をもとに、罫線・色・Excel数式（=数量×単価、=SUM）付きの見積書(.xlsx)を"
    "その場で生成してダウンロードできます。ダウンロード後もExcel上で数式が生きています。"
)

st.subheader("① 基本情報")
col1, col2 = st.columns(2)
with col1:
    customer_name = st.text_input("宛先（会社名）", value="株式会社サンプル")
with col2:
    subject = st.text_input("件名", value="クラウド導入支援 御見積")

st.subheader("② 明細行を編集")
default_items = pd.DataFrame(
    [
        {"品名": "クラウド基盤構築支援", "数量": 1, "単価": 800000},
        {"品名": "Streamlitアプリ開発", "数量": 1, "単価": 500000},
        {"品名": "保守サポート（月額×3）", "数量": 3, "単価": 50000},
    ]
)
items_df = st.data_editor(default_items, num_rows="dynamic", use_container_width=True, key="quote_items")

generate = st.button("📗 Excelを生成", type="primary")

if generate:
    items_df = items_df.dropna(subset=["品名"])
    if items_df.empty:
        st.error("明細行を1件以上入力してください。")
    else:
        items_df = items_df.copy()
        items_df[["数量", "単価"]] = items_df[["数量", "単価"]].fillna(0)

        wb = Workbook()
        ws = wb.active
        ws.title = "見積書"

        header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        title_font = Font(size=18, bold=True, color="1E3A8A")
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # タイトル
        ws.merge_cells("A1:D1")
        ws["A1"] = "御見積書"
        ws["A1"].font = title_font
        ws["A1"].alignment = Alignment(horizontal="center")

        ws["A3"] = "宛先:"
        ws["B3"] = f"{customer_name} 御中"
        ws["A4"] = "件名:"
        ws["B4"] = subject

        # 明細ヘッダー
        headers = ["品名", "数量", "単価", "金額"]
        header_row = 6
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=header_row, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center")

        # 明細データ（金額はExcel数式で計算）
        start_row = header_row + 1
        for i, row in enumerate(items_df.itertuples(index=False)):
            r = start_row + i
            ws.cell(row=r, column=1, value=row[0]).border = thin_border

            qty_cell = ws.cell(row=r, column=2, value=row[1])
            qty_cell.border = thin_border
            qty_cell.alignment = Alignment(horizontal="right")

            price_cell = ws.cell(row=r, column=3, value=row[2])
            price_cell.border = thin_border
            price_cell.number_format = "#,##0"

            amount_cell = ws.cell(row=r, column=4, value=f"=B{r}*C{r}")
            amount_cell.border = thin_border
            amount_cell.number_format = "#,##0"

        end_row = start_row + len(items_df) - 1
        total_row = end_row + 1

        label_cell = ws.cell(row=total_row, column=3, value="合計")
        label_cell.font = Font(bold=True)
        label_cell.alignment = Alignment(horizontal="right")
        label_cell.border = thin_border

        total_cell = ws.cell(row=total_row, column=4, value=f"=SUM(D{start_row}:D{end_row})")
        total_cell.font = Font(bold=True)
        total_cell.number_format = "#,##0"
        total_cell.border = thin_border

        # 列幅調整
        for i, w in enumerate([32, 10, 14, 14], start=1):
            ws.column_dimensions[get_column_letter(i)].width = w

        buf = io.BytesIO()
        wb.save(buf)
        excel_bytes = buf.getvalue()

        st.success("✅ Excelファイルを生成しました。")
        st.download_button(
            "💾 Excelファイルをダウンロード",
            data=excel_bytes,
            file_name=f"見積書_{customer_name}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )
